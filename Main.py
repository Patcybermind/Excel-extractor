import openpyxl as xl
#imports openpyxl

wb = xl.load_workbook("input.xlsx")
sheet = wb["input"]
#opens this sheet

print("Y: " + str(sheet.max_row))
print("X: " + str(sheet.max_column))
# just to clarify
# row = y
# column = x
x_output = ""
with open('output.txt', 'r+') as txt:
    # r = read, r+ = read and write, w = write
    # txt.write("test")
    for x in range(1, sheet.max_column + 3):
        for y in range(1, sheet.max_row + -1):
            cell = sheet.cell(x, y)
            if cell.value == None:
                x_output += "empty | "
            else:
                x_output += str(cell.value)  + " | "
        print(x_output)
        txt.write(x_output + "\n")
        x_output = ""